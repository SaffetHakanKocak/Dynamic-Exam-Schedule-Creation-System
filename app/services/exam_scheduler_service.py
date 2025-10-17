# app/services/exam_scheduler_service.py
import random
import datetime
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from app.db import fetchall, execute


class ExamSchedulerService:
    """Sınav Programı Oluşturma Servisi — Nihai Optimize Sürüm (tam kapsam + hata yönetimi)"""

    def __init__(self):
        self.errors = []
        self.generated_plan = []

    # ============================================================
    # 🔹 ANA METOD
    # ============================================================
    def generate_schedule(
        self,
        department_id: int,
        included_courses: list[str],
        start_date: str,
        end_date: str,
        holidays: list[str],
        exam_type: str,
        default_duration: int,
        gap_duration: int,
        no_overlap: bool,
        custom_durations: dict[str, int] | None = None
    ):
        self.errors.clear()
        self.generated_plan.clear()

        # 1️⃣ Tarih aralığı oluştur
        workdays = self._build_workdays(start_date, end_date, holidays)
        if not workdays:
            self.errors.append("Sınav yapılabilecek gün bulunamadı (tatil dışı gün yok)!")
            return []

        # 2️⃣ Dersleri getir
        courses = self._load_courses_and_students(department_id, included_courses)
        if not courses:
            self.errors.append("Hiç ders bulunamadı!")
            return []

        # 3️⃣ Derslikleri getir
        rooms = self._load_classrooms(department_id)
        if not rooms:
            self.errors.append("Bu bölüme ait derslik bulunamadı!")
            return []

        # ------------------------------------------------------------
        # ⚙️ Fizibilite Kontrolü
        # ------------------------------------------------------------
        days_count = len(workdays)
        slots_per_day = int((17 - 10) * 60 / (default_duration + gap_duration))  # 10:00–17:00 arası
        available_slots = days_count * slots_per_day
        total_courses = len(courses)

        if no_overlap and total_courses > available_slots:
            self.errors.append(
                f"⚠️ 'Hiçbir sınav aynı anda olmasın' seçeneği aktif.\n"
                f"{days_count} gün × {slots_per_day} slot = {available_slots} sınav planlanabilir.\n"
                f"Sistemde {total_courses} sınav var.\n"
                f"Tarih aralığını genişletin veya süre/bekleme süresini azaltın."
            )
            return []

        # 4️⃣ Sınıfa göre grupla
        grouped_by_class = self._group_by_class(courses)

        # 5️⃣ Her sınıfa gün ata
        class_day_map = self._assign_days_to_classes(grouped_by_class, workdays)

        # 6️⃣ Slotları oluştur
        slots_per_day = self._build_slots(default_duration, gap_duration)

        # 7️⃣ Yerleştirme
        placed_courses = []
        for cls_name, cls_courses in grouped_by_class.items():
            for i, course in enumerate(cls_courses):
                target_day = class_day_map[cls_name][i % len(class_day_map[cls_name])]
                placed = False
                last_error = None

                for slot_start in slots_per_day:
                    duration = (
                        custom_durations.get(course["code"], default_duration)
                        if custom_durations else default_duration
                    )
                    slot_end = (datetime.datetime.combine(target_day, slot_start)
                                + datetime.timedelta(minutes=duration)).time()

                    try:
                        # 1️⃣ Aynı slot dolu mu (no_overlap aktifse)
                        if no_overlap and self._overlaps_with_existing(
                            placed_courses, target_day, slot_start, slot_end
                        ):
                            raise Exception("Zaman çakışması")

                        # 2️⃣ Öğrenci çakışması kontrolü
                        if self._has_student_conflict(
                            course["id"], placed_courses, target_day, slot_start, slot_end
                        ):
                            raise Exception("Öğrenci çakışması")

                        # 3️⃣ Derslik ataması (kapasite ve slot uygunluğu)
                        assigned_rooms = self._assign_rooms(
                            needed=course["student_count"],
                            rooms=rooms,
                            placed=placed_courses,
                            target_day=target_day,
                            slot=f"{slot_start.strftime('%H:%M')} - {slot_end.strftime('%H:%M')}",
                            gap_min=gap_duration
                        )

                        if not assigned_rooms:
                            raise Exception("Derslik kapasitesi yetersiz")

                        # ✅ Başarılı yerleştirme
                        placed_courses.append({
                            "date": target_day.strftime("%d.%m.%Y"),
                            "slot": f"{slot_start.strftime('%H:%M')} - {slot_end.strftime('%H:%M')}",
                            "course": course,
                            "rooms": assigned_rooms,
                            "duration": duration,
                            "type": exam_type
                        })
                        placed = True
                        break

                    except Exception as e:
                        last_error = f"{course['code']} yerleştirilemedi ({str(e)})."
                        continue

                if not placed and last_error:
                    self.errors.append(last_error)

        # 8️⃣ DB'ye kaydet
        if not self.errors and placed_courses:
            self._persist_to_database(
                placed_courses, exam_type, start_date, end_date, default_duration, gap_duration
            )

        if self.errors:
            return []

        self.generated_plan = self._format_plan(placed_courses)
        return self.generated_plan

    # ============================================================
    # 🧩 Yardımcı Metodlar
    # ============================================================
    def _build_workdays(self, start_str, end_str, holidays):
        start = datetime.datetime.strptime(start_str, "%d.%m.%Y").date()
        end = datetime.datetime.strptime(end_str, "%d.%m.%Y").date()
        days = []
        for i in range((end - start).days + 1):
            d = start + datetime.timedelta(days=i)
            if d.strftime("%a").upper()[:3] in [h[:3].upper() for h in holidays]:
                continue
            days.append(d)
        return days

    def _build_slots(self, duration_min, gap_min):
        start_time = datetime.time(10, 0)
        slots = []
        current = datetime.datetime.combine(datetime.date.today(), start_time)
        end_limit = datetime.time(17, 0)
        while current.time() < end_limit:
            slots.append(current.time())
            current += datetime.timedelta(minutes=duration_min + gap_min)
        return slots

    def _load_courses_and_students(self, dept_id, included_codes):
        q = """
        SELECT c.id, c.code, c.name, c.instructor_name, c.class_name,
               COUNT(DISTINCT e.student_id) AS student_count
        FROM courses c
        LEFT JOIN enrollments e ON e.course_id = c.id
        WHERE c.department_id = %s
        """
        params = [dept_id]
        if included_codes:
            q += " AND c.code IN (" + ",".join(["%s"] * len(included_codes)) + ")"
            params += included_codes
        q += " GROUP BY c.id ORDER BY c.class_name, c.code"
        return fetchall(q, tuple(params))

    def _load_classrooms(self, dept_id):
        return fetchall("""
            SELECT id, code, capacity
            FROM classrooms
            WHERE department_id = %s
            ORDER BY capacity DESC
        """, (dept_id,))

    def _group_by_class(self, courses):
        grouped = {}
        for c in courses:
            cls = c["class_name"] or "Bilinmeyen"
            grouped.setdefault(cls, []).append(c)
        return grouped

    def _assign_days_to_classes(self, grouped, workdays):
        day_map = {}
        for cls_name, courses in grouped.items():
            shuffled_days = workdays.copy()
            random.shuffle(shuffled_days)
            day_map[cls_name] = shuffled_days[:len(courses)]
        return day_map

    def _overlaps_with_existing(self, placed, date, new_start, new_end):
        new_start_dt = datetime.datetime.combine(date, new_start)
        new_end_dt = datetime.datetime.combine(date, new_end)
        for p in placed:
            if p["date"] != date.strftime("%d.%m.%Y"):
                continue
            start_str, end_str = [x.strip() for x in p["slot"].split("-")]
            existing_start = datetime.datetime.strptime(start_str, "%H:%M")
            existing_end = datetime.datetime.strptime(end_str, "%H:%M")
            if existing_start < new_end_dt and new_start_dt < existing_end:
                return True
        return False

    def _has_student_conflict(self, course_id, placed, date, slot_start, slot_end):
        overlapping = []
        for p in placed:
            if p["date"] != date.strftime("%d.%m.%Y"):
                continue
            start_str, end_str = [x.strip() for x in p["slot"].split("-")]
            existing_start = datetime.datetime.strptime(start_str, "%H:%M")
            existing_end = datetime.datetime.strptime(end_str, "%H:%M")
            if existing_start.time() < slot_end and slot_start < existing_end.time():
                overlapping.append(p["course"]["id"])
        if not overlapping:
            return False
        q = f"""
        SELECT COUNT(*) AS c FROM enrollments e1
        JOIN enrollments e2 ON e1.student_id = e2.student_id
        WHERE e1.course_id = %s AND e2.course_id IN ({','.join(['%s']*len(overlapping))})
        """
        res = fetchall(q, tuple([course_id] + overlapping))
        return res[0]["c"] > 0 if res else False

    # ============================================================
    # 🧮 DERSLİK ATAMASI
    # ============================================================
    def _assign_rooms(self, needed, rooms, placed, target_day, slot, gap_min=15):
        assigned = []
        total_capacity = 0
        sorted_rooms = sorted(rooms, key=lambda r: r["capacity"], reverse=True)

        start_str, end_str = [x.strip() for x in slot.split("-")]
        slot_start = datetime.datetime.strptime(start_str, "%H:%M")
        slot_end = datetime.datetime.strptime(end_str, "%H:%M")
        slot_end_with_gap = slot_end + datetime.timedelta(minutes=gap_min)

        for room in sorted_rooms:
            conflict = False
            for p in placed:
                if p["date"] != target_day.strftime("%d.%m.%Y"):
                    continue
                for used in p["rooms"]:
                    if used["code"] != room["code"]:
                        continue
                    p_start, p_end = [datetime.datetime.strptime(x.strip(), "%H:%M") for x in p["slot"].split("-")]
                    p_end_with_gap = p_end + datetime.timedelta(minutes=gap_min)
                    if not (slot_end_with_gap <= p_start or slot_start >= p_end_with_gap):
                        conflict = True
                        break
                if conflict:
                    break
            if conflict:
                continue

            assigned.append(room)
            total_capacity += room["capacity"]
            if total_capacity >= needed:
                return assigned
        return []

    # ============================================================
    # 💾 VERİTABANI KAYDI
    # ============================================================
    def _persist_to_database(self, placed_courses, exam_type, start_date, end_date, duration, gap):
        """Oluşturulan sınav programını veritabanına güvenli şekilde kaydeder."""
        start_dt = datetime.datetime.strptime(start_date, "%d.%m.%Y").date()
        end_dt = datetime.datetime.strptime(end_date, "%d.%m.%Y").date()

        term_id = execute("""
            INSERT INTO exam_terms (name, date_start, date_end, default_duration_min, min_gap_min)
            VALUES (%s, %s, %s, %s, %s)
        """, (exam_type, start_dt, end_dt, duration, gap))

        slot_map = {}
        for p in placed_courses:
            date_str = p["date"]
            start_str, end_str = [x.strip() for x in p["slot"].split("-")]
            start_dt_full = datetime.datetime.strptime(f"{date_str} {start_str}", "%d.%m.%Y %H:%M")
            end_dt_full = datetime.datetime.strptime(f"{date_str} {end_str}", "%d.%m.%Y %H:%M")
            key = (start_dt_full, end_dt_full)

            if key not in slot_map:
                slot_id = execute("""
                    INSERT INTO timeslots (exam_term_id, starts_at, ends_at)
                    VALUES (%s, %s, %s)
                """, (term_id, start_dt_full, end_dt_full))
                slot_map[key] = slot_id

        for p in placed_courses:
            course_id = p["course"]["id"]
            date_str = p["date"]
            start_str, end_str = [x.strip() for x in p["slot"].split("-")]
            start_dt_full = datetime.datetime.strptime(f"{date_str} {start_str}", "%d.%m.%Y %H:%M")
            end_dt_full = datetime.datetime.strptime(f"{date_str} {end_str}", "%d.%m.%Y %H:%M")
            key = (start_dt_full, end_dt_full)
            slot_id = slot_map.get(key)
            if not slot_id:
                continue

            exam_id = execute("""
                INSERT INTO exams (course_id, exam_term_id, timeslot_id, status)
                VALUES (%s, %s, %s, 'PLANNED')
            """, (course_id, term_id, slot_id))

            for r in p["rooms"]:
                execute("""
                    INSERT INTO exam_rooms (exam_id, classroom_id)
                    VALUES (%s, %s)
                """, (exam_id, r["id"]))

    # ============================================================
    def _format_plan(self, placed_courses):
        return [{
            "Tarih": p["date"],
            "Saat": p["slot"],
            "Ders": f"{p['course']['code']} — {p['course']['name']}",
            "Öğretim Elemanı": p["course"]["instructor_name"],
            "Derslikler": ", ".join([r["code"] for r in p["rooms"]]),
            "Tür": p["type"],
            "Süre (dk)": p["duration"]
        } for p in placed_courses]

    # ============================================================
    def export_to_excel(self, plan, filename):
        # 🔹 Tarih sırasına göre düzenle
        sorted_plan = sorted(
            plan,
            key=lambda x: (
                datetime.datetime.strptime(x["Tarih"], "%d.%m.%Y"),
                x["Saat"].split(" - ")[0]
            )
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sınav Programı"

        headers = ["Tarih", "Saat", "Ders", "Öğretim Elemanı", "Derslikler", "Tür", "Süre (dk)"]
        ws.append(headers)

        # 🔹 Gün ayracı için renkler
        day_fill = PatternFill(start_color="E8F0FE", fill_type="solid")

        prev_date = None
        for row in sorted_plan:
            ws.append([row[h] for h in headers])
            current_row = ws.max_row
            if row["Tarih"] != prev_date:
                for cell in ws[current_row]:
                    cell.fill = day_fill
                prev_date = row["Tarih"]

        # 🔹 Başlık biçimlendirme
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", fill_type="solid")
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

        # 🔹 Hizalama
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        wb.save(filename)
        return True
